import openpyxl
from difflib import SequenceMatcher

def goThroughExcel(l):
    #1. Read the Excel, get the number of rows m_row
    path = "test.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    print("We have %d rows in the Excel."%m_row)
    #2. Use a for loop to go through each row,
    #   calculate the similar rate of each company's name with compang name list

    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        print (cell_obj.value)
        similar_rate_result = []
        for j in range(0, len(l)):
            similar_rate_result.append(similar(cell_obj.value, l[j]))
        print (cell_obj, similar_rate_result)
        print (similar_rate_result.index(max(similar_rate_result)))
        target_name = l[similar_rate_result.index(max(similar_rate_result))]
        sheet_obj.cell(row = i, column = 2).value = target_name
        sheet_obj.cell(row=i, column=3).value = max(similar_rate_result)
    wb_obj.save("test.xlsx")

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


if __name__ == '__main__':
    company_name_list = ['Apple','Bee','PIG','Giao']
    print(similar('D&I', 'doe and Ingalls'))
    goThroughExcel(company_name_list)


